#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import math
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
A = ROOT / 'analysis'
D = ROOT / 'deliverables'
F = ROOT / 'frozen'

checks: list[dict[str, object]] = []
warnings: list[str] = []


def record(name: str, passed: bool, detail: str = '') -> None:
    checks.append({'check': name, 'passed': bool(passed), 'detail': detail})


def same(a: float, b: float, tol: float = 0.75) -> bool:
    return bool(np.isfinite(a) and np.isfinite(b) and abs(float(a) - float(b)) <= tol)


def sha(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def verify_sidecar(path: Path) -> None:
    side = Path(str(path) + '.sha256')
    record(f'Frozen file exists: {path.name}', path.exists(), str(path.relative_to(ROOT)))
    record(f'Frozen hash sidecar exists: {path.name}', side.exists(), str(side.relative_to(ROOT)))
    if path.exists() and side.exists():
        expected = side.read_text(encoding='utf-8').strip().split()[0]
        actual = sha(path)
        record(f'Frozen checksum: {path.name}', actual == expected, f'expected={expected}; actual={actual}')


# Required deliverables and frozen inputs.
required = [
    D/'rp_ceo_expanded_rebenchmark_report.pdf',
    D/'rp_ceo_expanded_rebenchmark_report.md',
    D/'rp_ceo_expanded_benchmark_workbook.xlsx',
    D/'expanded_reference_set.csv',
    D/'form990_evidence.csv',
    D/'job_ad_evidence.csv',
    D/'peer_inclusion_exclusion_log.csv',
    D/'rp_public_profile.csv',
    D/'source_manifest.csv',
    A/'stats.json',
    A/'summary_statistics.csv',
    ROOT/'data/form990_evidence_working.csv',
    ROOT/'data/job_ad_evidence_expanded.csv',
    ROOT/'data/original_source_manifest.csv',
]
for p in required:
    record(f'Required file: {p.relative_to(ROOT)}', p.exists() and p.stat().st_size > 0, f'bytes={p.stat().st_size if p.exists() else 0}')

for p in [
    F/'extension2_protocol_amendment.md',
    F/'extension2_new_candidates_precomp.csv',
    F/'combined_peer_universe_precomp.csv',
    F/'extension3_protocol_amendment.md',
    F/'extension3_new_candidates_precomp.csv',
    F/'combined_peer_universe_3wave_precomp.csv',
    F/'legacy_original_peer_bridge_precomp.csv',
]:
    verify_sidecar(p)

stats = json.loads((A/'stats.json').read_text(encoding='utf-8'))
ref = pd.read_csv(D/'expanded_reference_set.csv')
filings = pd.read_csv(D/'form990_evidence.csv')
primary = filings[filings['primary_eligible'].astype(str).str.lower().eq('true')].copy()
clean = primary[primary['structurally_clean'].astype(str).str.lower().eq('true')].copy()
jobs = pd.read_csv(D/'job_ad_evidence.csv')
qjobs = jobs[jobs['included_in_quantitative_analysis'].eq('yes')].copy()
peerlog = pd.read_csv(D/'peer_inclusion_exclusion_log.csv')
manifest = pd.read_csv(D/'source_manifest.csv')

# Universe and reference-set integrity.
record('Third-wave frozen expansion universe has 450 organizations', len(pd.read_csv(F/'combined_peer_universe_3wave_precomp.csv')) == 450)
record('Second-wave candidate freeze has 145 organizations', len(pd.read_csv(F/'extension2_new_candidates_precomp.csv')) == 145)
record('Third-wave candidate freeze has 145 organizations', len(pd.read_csv(F/'extension3_new_candidates_precomp.csv')) == 145)
record('Documented peer log has 452 unique organizations', len(peerlog) == 452 and peerlog.organization.nunique() == 452)
record('Selected reference set has 144 unique organizations', len(ref) == 144 and ref.organization.nunique() == 144)
record('Selected tiers total correctly', ref.reference_tier.value_counts().to_dict() == {'A':79, 'B':39, 'C':26}, str(ref.reference_tier.value_counts().to_dict()))
record('No selected reference organization was marked pay-seen before freeze', not peerlog.loc[peerlog.final_status.str.startswith('selected', na=False), 'compensation_seen_before_freeze'].astype(str).str.lower().eq('yes').any())
record('Stats selected-reference count agrees', stats['selected_reference_n'] == len(ref), f"stats={stats['selected_reference_n']}; csv={len(ref)}")

# Filing integrity.
required_primary_cols = [
    'organization','ein','irs_object_id','tax_period_begin','tax_period_end','filing_date',
    'ceo_name','ceo_title','compensation_calendar_year','part_vii_org','part_vii_related',
    'part_vii_other','cash_proxy','total_proxy','revenue','expenses','official_irs_url',
    'propublica_url','source_id','extraction_location'
]
for col in required_primary_cols:
    ok = col in primary and primary[col].notna().all() and primary[col].astype(str).str.strip().ne('').all()
    record(f'Primary filing required field complete: {col}', ok, f'missing={int(primary[col].isna().sum()) if col in primary else "column absent"}')
record('Primary filing organizations unique', primary.organization.nunique() == len(primary), f'n={len(primary)}; unique={primary.organization.nunique()}')
record('Primary filing observation count is 116', len(primary) == 116)
record('Structurally clean filing count is 79', len(clean) == 79)
record('At least 100 primary-use filing observations', len(primary) >= 100)
record('Cash proxy arithmetic', np.allclose(primary.cash_proxy, primary.part_vii_org.fillna(0) + primary.part_vii_related.fillna(0), atol=0.01))
record('Total proxy arithmetic', np.allclose(primary.total_proxy, primary.cash_proxy.fillna(0) + primary.part_vii_other.fillna(0), atol=0.01))
record('CPI-normalized cash arithmetic', np.allclose(primary.cash_proxy_jul2026, primary.cash_proxy * primary.cpi_factor, atol=0.02))
record('CPI-normalized total arithmetic', np.allclose(primary.total_proxy_jul2026, primary.total_proxy * primary.cpi_factor, atol=0.02))
filing_dates = pd.to_datetime(primary.filing_date, errors='coerce')
record('Primary filing dates parse', filing_dates.notna().all())
record('No primary filing after cutoff', filing_dates.le(pd.Timestamp('2026-08-17')).all())
record('No exact Schedule J base silently imputed', primary.schedule_j_base.notna().sum() == 0 and stats['schedule_j_exact_base_n'] == 0)
record('Stats primary count agrees', stats['primary_incumbent_n'] == len(primary))
record('Stats clean count agrees', stats['structurally_clean_incumbent_n'] == len(clean))
record('Known staff availability disclosed', int(primary.employee_count.notna().sum()) == 43, f'known={primary.employee_count.notna().sum()}; missing={primary.employee_count.isna().sum()}')
warnings.append(f'Employee count is unavailable for {int(primary.employee_count.isna().sum())} of {len(primary)} primary observations; scale analyses therefore report both expense-only and joint known-staff views.')

# Reproduce headline filing statistics.
def med(df: pd.DataFrame, col: str) -> float:
    return float(pd.to_numeric(df[col], errors='coerce').median())

samples = {
    'all_primary_eligible': primary,
    'structurally_clean': clean,
    'tier_A': primary[primary.tier_group.eq('A')],
    'tier_A_B': primary[primary.tier_group.isin(['A','B'])],
    'expense_and_available_staff_match': primary[primary.close_scale.astype(str).str.lower().eq('true')],
    'joint_expense_and_known_staff_match': primary[primary.joint_scale_known_staff.astype(str).str.lower().eq('true')],
    'EA_core_or_adjacent': primary[primary.ea_affinity.ne('functional-only')],
    'functional_only': primary[primary.ea_affinity.eq('functional-only')],
}
for name, df in samples.items():
    expected = stats['form990'][name]
    record(f'Statistic n: {name}', len(df) == expected['n'], f'csv={len(df)}; stats={expected["n"]}')
    record(f'Cash median: {name}', same(med(df,'cash_proxy_jul2026'), expected['cash']['median']))
    record(f'Total median: {name}', same(med(df,'total_proxy_jul2026'), expected['total']['median']))

# Leave-one-out close-scale calculation.
close = samples['expense_and_available_staff_match']
loo_cash=[]; loo_total=[]
for org in close.organization:
    d=close[close.organization.ne(org)]
    loo_cash.append(med(d,'cash_proxy_jul2026'))
    loo_total.append(med(d,'total_proxy_jul2026'))
loo=stats['leave_one_out_close']
record('Leave-one-out close-scale n', len(close) == loo['n'])
record('Leave-one-out cash minimum', same(min(loo_cash), loo['cash_median_min']))
record('Leave-one-out cash maximum', same(max(loo_cash), loo['cash_median_max']))
record('Leave-one-out total minimum', same(min(loo_total), loo['total_median_min']))
record('Leave-one-out total maximum', same(max(loo_total), loo['total_median_max']))

# Job-ad integrity and statistics.
for col in ['organization','role_title','posting_date','salary_min','salary_max','location','original_url','retrieved_at','source_id','archive_path']:
    ok = qjobs[col].notna().all() and qjobs[col].astype(str).str.strip().ne('').all()
    record(f'Quant job-ad required field complete: {col}', ok)
record('Quantitative job-ad count is 15', len(qjobs) == 15)
record('Job-ad organizations unique', qjobs.organization.nunique() == len(qjobs))
record('Job-ad ranges valid', qjobs.salary_min.le(qjobs.salary_max).all())
record('Job-ad archive paths exist', all((ROOT/str(x)).exists() for x in qjobs.archive_path))
record('Stats job-ad count agrees', stats['job_ads_current_quantitative_n'] == len(qjobs))
for name, mask in {
    'strict_primary': qjobs.tier.eq('strict_primary'),
    'close_title_scale': qjobs.tier.isin(['strict_primary','expanded_primary_title','expanded_broad_functional']),
    'expanded_current': pd.Series(True,index=qjobs.index),
}.items():
    d=qjobs[mask]
    expected=stats['job_ads'][name]
    record(f'Job-ad sample n: {name}', len(d)==expected['n'])
    record(f'Job-ad midpoint median: {name}', same(float(d.adjusted_midpoint_jul2026.median()), expected['median_midpoint']))

# Source manifest checks.
record('Source manifest IDs unique', manifest.source_id.nunique() == len(manifest))
missing=[]; bad_bytes=[]; bad_hash=[]
for _, r in manifest.iterrows():
    p=ROOT/str(r.local_archive_path)
    if not p.exists():
        missing.append(str(r.local_archive_path)); continue
    if p.stat().st_size != int(r.byte_length): bad_bytes.append(str(r.local_archive_path))
    if sha(p) != str(r.sha256): bad_hash.append(str(r.local_archive_path))
record('All manifest archive paths exist', not missing, '; '.join(missing[:5]))
record('All manifest byte lengths match', not bad_bytes, '; '.join(bad_bytes[:5]))
record('All manifest SHA-256 values match', not bad_hash, '; '.join(bad_hash[:5]))
record('Every primary filing source appears in manifest', set(primary.source_id).issubset(set(manifest.source_id)))
record('Every quantitative job-ad source appears in manifest', set(qjobs.source_id).issubset(set(manifest.source_id)))
warnings.append('The archive preserves filing-derived snapshots and canonical IRS identifiers, not a complete source-native XML/PDF copy for every filing.')

# Report checks.
report_md=(D/'rp_ceo_expanded_rebenchmark_report.md').read_text(encoding='utf-8')
for needle in [
    '144 selected organizations','116 primary-use incumbent observations','79 structurally clean',
    'Advertise $250K-$340K','Successful incumbent base $290K-$350K',
    '$340K-$435K','Exact Schedule J base','not blind to the original benchmark',
    'd84b6c772f8017bd4ddb8a4327012aabe40b6cd25198552558b4018f6c32e4c3'
]:
    record(f'Report contains: {needle}', needle in report_md)
reader=PdfReader(str(D/'rp_ceo_expanded_rebenchmark_report.pdf'))
record('Report PDF has expected substantive length', 15 <= len(reader.pages) <= 30, f'pages={len(reader.pages)}')
text='\n'.join((p.extract_text() or '') for p in reader.pages)
record('PDF contains headline posting range', '$250K-$340K' in text or '$250K-$340K base' in text)
record('PDF contains primary observation count', '116 primary-use' in text)

# Workbook checks.
wb=load_workbook(D/'rp_ceo_expanded_benchmark_workbook.xlsx', data_only=False, read_only=False)
required_sheets={'Summary','Reference Set','Primary 990','All 990','Quant Job Ads','All Job Ads','Sensitivities','Peer Log','Source Manifest','Protocol & Notes'}
record('Workbook contains required sheets', required_sheets.issubset(set(wb.sheetnames)), str(wb.sheetnames))
formula_cells=[]; error_literals=[]
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v=cell.value
            if isinstance(v,str) and v.startswith('='): formula_cells.append((ws.title,cell.coordinate,v))
            if isinstance(v,str) and any(e in v for e in ['#REF!','#DIV/0!','#VALUE!','#NAME?','#N/A']): error_literals.append((ws.title,cell.coordinate,v))
record('Workbook contains formulas for derived values', len(formula_cells) >= 20, f'formulas={len(formula_cells)}')
record('Workbook contains no formula-error literals', not error_literals, str(error_literals[:5]))
record('Workbook reference-set rows agree', wb['Reference Set'].max_row - 4 == len(ref), f'workbook={wb["Reference Set"].max_row-4}; csv={len(ref)}')
record('Workbook primary filing rows agree', wb['Primary 990'].max_row - 4 == len(primary), f'workbook={wb["Primary 990"].max_row-4}; csv={len(primary)}')
record('Workbook quantitative job-ad rows agree', wb['Quant Job Ads'].max_row - 4 == len(qjobs), f'workbook={wb["Quant Job Ads"].max_row-4}; csv={len(qjobs)}')

# Recommendation bounds are intentionally analyst judgment, but must be internally ordered.
j=stats['analyst_judgment']
for name in ['external_posting','incumbent_base','recurring_total']:
    record(f'Analyst range ordered: {name}', j[name]['low'] < j[name]['center'] < j[name]['high'], str(j[name]))
record('Expected-hire zone within posting range', j['external_posting']['low'] <= j['external_posting']['expected_hire_low'] < j['external_posting']['expected_hire_high'] <= j['external_posting']['high'])

passed=sum(1 for c in checks if c['passed'])
failed=[c for c in checks if not c['passed']]
status='PASS WITH DOCUMENTED SOURCE-ARCHIVE AND STAFF-DATA LIMITATIONS' if not failed else 'FAIL'
lines=[
    'RP CEO EXPANDED REBENCHMARK - INDEPENDENT VALIDATION REPORT',
    f'Status: {status}',
    f'Checks passed: {passed}',
    f'Checks failed: {len(failed)}',
    f'Warnings: {len(warnings)}',
    '',
]
for c in checks:
    lines.append(f"[{'PASS' if c['passed'] else 'FAIL'}] {c['check']}" + (f" - {c['detail']}" if c['detail'] else ''))
lines += ['', 'WARNINGS'] + [f'- {w}' for w in warnings]
(A/'validation_report.txt').write_text('\n'.join(lines)+'\n',encoding='utf-8')
(A/'validation_results.json').write_text(json.dumps({'status':status,'passed':passed,'failed':failed,'warnings':warnings,'checks':checks},indent=2),encoding='utf-8')
print('\n'.join(lines[:8]))
if failed:
    for c in failed:
        print('FAILED:',c['check'],c['detail'],file=sys.stderr)
    raise SystemExit(1)
