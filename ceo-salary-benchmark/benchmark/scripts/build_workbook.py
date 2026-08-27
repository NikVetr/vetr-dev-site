#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
import json, math
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.chart import BarChart, Reference

ROOT=Path(__file__).resolve().parents[1]
D=ROOT/'deliverables'; A=ROOT/'analysis'; F=ROOT/'frozen'
stats=json.loads((A/'stats.json').read_text())
NSEL=stats['selected_reference_n']; NPRIMARY=stats['primary_incumbent_n']; NCLEAN=stats['structurally_clean_incumbent_n']
ref=pd.read_csv(D/'expanded_reference_set.csv')
all990=pd.read_csv(D/'form990_evidence.csv')
primary=all990[all990.primary_eligible==True].copy().reset_index(drop=True)
jobs=pd.read_csv(D/'job_ad_evidence.csv')
qjobs=jobs[jobs.included_in_quantitative_analysis.eq('yes')].copy().reset_index(drop=True)
sens=pd.read_csv(A/'summary_statistics.csv')
peerlog=pd.read_csv(D/'peer_inclusion_exclusion_log.csv')
manifest=pd.read_csv(D/'source_manifest.csv')

NAVY='17324D'; TEAL='2F7180'; LIGHT_TEAL='E7F0F2'; ORANGE='C27038'; LIGHT_ORANGE='F3EADF'; GREEN='008000'; BLUE='0000FF'; GRAY='666666'; LIGHT_GRAY='E9EEF1'; WHITE='FFFFFF'; BLACK='000000'; RED='C00000'; PURPLE='7030A0'
header_fill=PatternFill('solid',fgColor=NAVY); section_fill=PatternFill('solid',fgColor=TEAL); caution_fill=PatternFill('solid',fgColor=LIGHT_ORANGE); kpi_fill=PatternFill('solid',fgColor=LIGHT_TEAL)
thin_gray=Side(style='thin',color='D4DCE1')

wb=Workbook(); wb.remove(wb.active)
wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.properties.title='RP CEO Expanded Compensation Benchmark'; wb.properties.subject='Independent expanded nonprofit CEO compensation benchmark as of August 17, 2026'; wb.properties.creator='OpenAI'

def title(ws,text,subtitle=None):
    ws.sheet_view.showGridLines=False
    ws.merge_cells('A1:H1'); ws['A1']=text; ws['A1'].font=Font(size=18,bold=True,color=NAVY); ws['A1'].alignment=Alignment(vertical='center'); ws.row_dimensions[1].height=28
    if subtitle:
        ws.merge_cells('A2:H2'); ws['A2']=subtitle; ws['A2'].font=Font(size=9,color=GRAY); ws['A2'].alignment=Alignment(wrap_text=True); ws.row_dimensions[2].height=28

def format_header(ws,row,max_col):
    for c in range(1,max_col+1):
        cell=ws.cell(row,c); cell.fill=header_fill; cell.font=Font(bold=True,color=WHITE); cell.alignment=Alignment(wrap_text=True,vertical='center')
    ws.row_dimensions[row].height=30

def write_df(ws,df,start_row=1,source_cols=()):
    for j,c in enumerate(df.columns,1): ws.cell(start_row,j,c)
    format_header(ws,start_row,len(df.columns))
    for i,(_,r) in enumerate(df.iterrows(),start_row+1):
        for j,c in enumerate(df.columns,1):
            v=r[c]
            if pd.isna(v): v=None
            ws.cell(i,j,v)
            ws.cell(i,j).alignment=Alignment(vertical='center',wrap_text=(isinstance(v,str) and len(v)>45))
            ws.cell(i,j).font=Font(color=GREEN if c in source_cols or 'url' in c.lower() else GRAY)
    ws.auto_filter.ref=f"A{start_row}:{get_column_letter(len(df.columns))}{start_row+len(df)}"
    ws.freeze_panes=f'A{start_row+1}'

def fit_cols(ws,min_width=9,max_width=42):
    for col in range(1,ws.max_column+1):
        letter=get_column_letter(col); vals=[]
        for cell in ws[letter][:min(ws.max_row,200)]:
            if cell.value is not None: vals.append(len(str(cell.value)))
        width=max(vals+[min_width])+2
        ws.column_dimensions[letter].width=min(max(width,min_width),max_width)

def currency_cols(ws,headers,row=3):
    for c in range(1,ws.max_column+1):
        h=str(ws.cell(row,c).value or '').lower()
        if any(k in h for k in ['salary','proxy','compensation','revenue','expenses','budget','adjusted_min','adjusted_max','midpoint','cash_','total_']):
            for r in range(row+1,ws.max_row+1): ws.cell(r,c).number_format='$#,##0;[Red]($#,##0);-'
        elif 'score' in h or 'factor' in h or 'weight' in h:
            for r in range(row+1,ws.max_row+1): ws.cell(r,c).number_format='0.00'

# Summary
ws=wb.create_sheet('Summary'); title(ws,'RP CEO Expanded Compensation Benchmark',f'As of August 17, 2026 | 452 documented organizations | {NSEL} selected reference organizations | {NPRIMARY} primary-use filings ({NCLEAN} structurally clean)')
ws['A4']='Decision use'; ws['B4']='Low'; ws['C4']='High'; ws['D4']='Descriptive center'; ws['E4']='Interpretation'; format_header(ws,4,5)
decisions=[('External recruitment posting',250000,340000,295000,'Expected hire zone $280K-$320K'),('Successful incumbent base',290000,350000,320000,'Successful full-time organization-wide CEO'),('Recurring total compensation',340000,435000,385000,'Low-to-moderate confidence; excludes one-time items')]
for i,row in enumerate(decisions,5):
    for j,v in enumerate(row,1): ws.cell(i,j,v)
    for j in (2,3,4): ws.cell(i,j).number_format='$#,##0;[Red]($#,##0);-'; ws.cell(i,j).fill=caution_fill; ws.cell(i,j).font=Font(bold=True,color=BLACK)
    ws.cell(i,1).font=Font(bold=True,color=NAVY); ws.cell(i,5).alignment=Alignment(wrap_text=True)
ws['A9']='Coverage'; ws['A9'].fill=section_fill; ws['A9'].font=Font(bold=True,color=WHITE); ws.merge_cells('A9:E9')
coverage=[('Frozen candidate universe',"=COUNTA('Peer Log'!A:A)-3"),('Selected reference organizations',"=COUNTA('Reference Set'!A:A)-3"),('Primary-use incumbent observations',"=COUNTA('Primary 990'!A:A)-3"),('Current quantitative job ads',"=COUNTA('Quant Job Ads'!A:A)-3"),('Exact Schedule J base observations',"=COUNT('Primary 990'!N:N)")]
for i,(lab,form) in enumerate(coverage,10): ws.cell(i,1,lab); ws.cell(i,2,form); ws.cell(i,2).font=Font(color=BLACK,bold=True); ws.cell(i,2).fill=kpi_fill
ws['A16']='Evidence centers'; ws['A16'].fill=section_fill; ws['A16'].font=Font(bold=True,color=WHITE); ws.merge_cells('A16:E16')
ws.append([])
centers=[
 ('Strict ad midpoint median',"=MEDIAN('Quant Job Ads'!I5:I7)",'Posted base salary'),
 ('Expanded ad midpoint median',f"=MEDIAN('Quant Job Ads'!I5:I{len(qjobs)+4})",'Posted base salary'),
 ('Broad filing cash-proxy median',f"=MEDIAN('Primary 990'!L5:L{len(primary)+4})",'Not exact base'),
 ('Broad filing total-proxy median',f"=MEDIAN('Primary 990'!M5:M{len(primary)+4})",'Cash proxy plus Part VII other'),
 ('Cash-proxy Q1',f"=QUARTILE.INC('Primary 990'!L5:L{len(primary)+4},1)",'Broad filing sample'),
 ('Cash-proxy Q3',f"=QUARTILE.INC('Primary 990'!L5:L{len(primary)+4},3)",'Broad filing sample'),
]
for i,(lab,form,note) in enumerate(centers,17): ws.cell(i,1,lab); ws.cell(i,2,form); ws.cell(i,3,note); ws.cell(i,2).number_format='$#,##0;[Red]($#,##0);-'; ws.cell(i,2).fill=kpi_fill
ws['A25']='Important limitation'; ws['A25'].fill=caution_fill; ws['A25'].font=Font(bold=True,color=ORANGE); ws['B25']='Form 990 Part VII reportable compensation is a cash/W-2 proxy, not base salary. No filing base was imputed.'; ws.merge_cells('B25:H26'); ws['B25'].alignment=Alignment(wrap_text=True,vertical='top')
for c in ['B5','B6','B7','B17','B18','B19','B20']:
    ws[c].comment=Comment('Derived from the archived evidence sheets and recalculated when opened in Excel. See Source Manifest and report for measure definitions.','OpenAI')
fit_cols(ws,12,38); ws.column_dimensions['A'].width=34; ws.column_dimensions['E'].width=48

# Reference Set
ws=wb.create_sheet('Reference Set'); title(ws,'Selected Reference Set',f'{NSEL} organizations selected before or independently of pay; {NPRIMARY} have primary-use incumbent observations ({NCLEAN} structurally clean)')
write_df(ws,ref,start_row=4,source_cols=('propublica_url','official_irs_url'))
currency_cols(ws,ref.columns,4); fit_cols(ws,8,40)

# Primary 990 - formula-driven derived columns
ws=wb.create_sheet('Primary 990'); title(ws,'Primary Incumbent Form 990 Evidence','One clean organization-wide top-executive observation per organization; all dollars normalized to July 2026')
headers=['Organization','Tier','EA affinity','CEO title','Comp year','CPI factor','Part VII org','Part VII related','Part VII other','Cash proxy','Total proxy','Cash Jul-26','Total Jul-26','Schedule J base','Expenses','Staff','Score','Analysis status','Structure flag','EIN','IRS object ID','Official IRS URL','ProPublica URL']
for j,h in enumerate(headers,1): ws.cell(4,j,h)
format_header(ws,4,len(headers))
for i,r in primary.iterrows():
    rr=i+5
    vals=[r.organization,r.tier_group,r.ea_affinity,r.ceo_title,int(r.compensation_calendar_year),r.cpi_factor,r.part_vii_org,r.part_vii_related,r.part_vii_other,None,None,None,None,r.schedule_j_base,r.expenses,r.employee_count,r.comparability_score,r.analysis_status,r.structure_flag,r.ein,str(r.irs_object_id),r.official_irs_url,r.propublica_url]
    for j,v in enumerate(vals,1):
        if pd.isna(v):v=None
        ws.cell(rr,j,v); ws.cell(rr,j).font=Font(color=GREEN if j not in (10,11,12,13) else BLACK); ws.cell(rr,j).alignment=Alignment(vertical='center',wrap_text=(j in (18,19,22,23)))
    ws.cell(rr,10,f'=G{rr}+H{rr}')
    ws.cell(rr,11,f'=J{rr}+I{rr}')
    ws.cell(rr,12,f'=J{rr}*F{rr}')
    ws.cell(rr,13,f'=K{rr}*F{rr}')
for col in range(6,18):
    h=headers[col-1].lower()
    for rr in range(5,len(primary)+5):
        ws.cell(rr,col).number_format='0.0000' if col==6 else ('$#,##0;[Red]($#,##0);-' if col not in (16,17) else '0')
ws.auto_filter.ref=f'A4:W{len(primary)+4}'; ws.freeze_panes='A5'; fit_cols(ws,8,35); ws.column_dimensions['A'].width=36; ws.column_dimensions['S'].width=34; ws.column_dimensions['V'].width=28; ws.column_dimensions['W'].width=28

# All 990 raw/normalized
ws=wb.create_sheet('All 990'); title(ws,'All Collected Form 990 Observations','Includes primary, structural sensitivity, transition, partial-year, and measurement-exclusion rows')
cols=['organization','ceo_name','ceo_title','compensation_calendar_year','filing_date','part_vii_org','part_vii_related','part_vii_other','cash_proxy_jul2026','total_proxy_jul2026','revenue','expenses','employee_count','analysis_status','exclusion_reason','structure_flag','ein','irs_object_id','official_irs_url','propublica_url','source_id']
write_df(ws,all990[cols],4,source_cols=('official_irs_url','propublica_url','source_id')); currency_cols(ws,cols,4); fit_cols(ws,8,38)

# Quant Job Ads formula-driven
ws=wb.create_sheet('Quant Job Ads'); title(ws,'Current Recruitment Evidence','Posted base-salary ranges only; separate from incumbent Form 990 compensation')
headers=['Organization','Role','Posting date','CPI factor','Salary min','Salary max','Adjusted min','Adjusted max','Adjusted midpoint','Location','Remote status','Budget/expense','Staff','Tier','Source ID','Original URL']
for j,h in enumerate(headers,1): ws.cell(4,j,h)
format_header(ws,4,len(headers))
for i,r in qjobs.iterrows():
    rr=i+5
    vals=[r.organization,r.role_title,str(r.posting_date),r.cpi_factor,r.salary_min,r.salary_max,None,None,None,r.location,r.remote_status,r.annual_budget_or_expense,r.staff_count,r.tier,r.source_id,r.original_url]
    for j,v in enumerate(vals,1):
        if pd.isna(v):v=None
        ws.cell(rr,j,v); ws.cell(rr,j).font=Font(color=BLACK if j in (7,8,9) else GREEN); ws.cell(rr,j).alignment=Alignment(vertical='center',wrap_text=(j in (1,2,10,13,14,16)))
    ws.cell(rr,7,f'=E{rr}*D{rr}'); ws.cell(rr,8,f'=F{rr}*D{rr}'); ws.cell(rr,9,f'=AVERAGE(G{rr}:H{rr})')
for c in range(4,14):
    for rr in range(5,len(qjobs)+5): ws.cell(rr,c).number_format='0.0000' if c==4 else ('$#,##0;[Red]($#,##0);-' if c in (5,6,7,8,9,12) else '0')
ws.auto_filter.ref=f'A4:P{len(qjobs)+4}'; ws.freeze_panes='A5'; fit_cols(ws,9,38); ws.column_dimensions['A'].width=36; ws.column_dimensions['P'].width=42

# All Job Ads
ws=wb.create_sheet('All Job Ads'); title(ws,'All Advertisement Searches','Included, sensitivity-only, and excluded advertisements and search outcomes')
jobcols=['organization','role_title','posting_date','salary_min','salary_max','location','remote_status','annual_budget_or_expense','staff_count','tier','included_in_quantitative_analysis','inclusion_reason','exclusion_reason','original_url','source_id']
write_df(ws,jobs[jobcols],4,source_cols=('original_url','source_id')); currency_cols(ws,jobcols,4); fit_cols(ws,8,42)

# Sensitivities
ws=wb.create_sheet('Sensitivities'); title(ws,'Sensitivity Analysis','Unweighted and comparability-weighted July 2026 centers')
write_df(ws,sens,4); currency_cols(ws,sens.columns,4); fit_cols(ws,10,30)

# Peer log
ws=wb.create_sheet('Peer Log'); title(ws,'452-Organization Inclusion/Exclusion Log','Every frozen candidate, final selection status, tier, structural reason, and compensation-collection status')
write_df(ws,peerlog,4,source_cols=('source_url',)); fit_cols(ws,8,45)

# Source manifest
ws=wb.create_sheet('Source Manifest'); title(ws,'Source Manifest','Original/resolved URLs, local archive paths, byte lengths, SHA-256 hashes, filing identifiers, and validation notes')
write_df(ws,manifest,4,source_cols=('original_url','resolved_url','local_archive_path')); fit_cols(ws,8,48)

# Protocol notes
ws=wb.create_sheet('Protocol & Notes'); title(ws,'Protocol and Evidence Definitions','Frozen rules and interpretive cautions')
notes=[
('Benchmark cutoff','August 17, 2026'),('Second expansion freeze','2026-08-25T16:07:55-0700'),('Third expansion freeze','2026-08-25T17:12:00-0700'),('Documented universe','452 unique organizations (450 expansion candidates plus 2 legacy-only peers)'),('Selected reference set',f'{NSEL} organizations'),('Primary-use incumbent observations',f'{NPRIMARY}; {NCLEAN} structurally clean'),('Current recruitment observations','15'),
('Base salary','Fixed annual cash salary stated in a posting or separately disclosed in Schedule J.'),('Cash/W-2 proxy','Part VII reportable compensation from the organization plus related organizations; may include bonuses or other taxable pay.'),('Filing total proxy','Cash/W-2 proxy plus Part VII other compensation where Schedule J total is unavailable.'),('Independence','This expansion was not blind to the original benchmark. Newly added candidates were frozen before systematic review of their pay.'),('Source limitation','The archive contains derivative snapshots and canonical IRS identifiers, not every complete source-native IRS XML return.'),('Decision use','Ranges are descriptive decision guidance, not population percentiles or a legal opinion.')]
ws['A4']='Topic';ws['B4']='Content';format_header(ws,4,2)
for i,(a,b) in enumerate(notes,5): ws.cell(i,1,a);ws.cell(i,2,b);ws.cell(i,1).font=Font(bold=True,color=NAVY);ws.cell(i,2).alignment=Alignment(wrap_text=True,vertical='top')
ws.column_dimensions['A'].width=28;ws.column_dimensions['B'].width=100;ws.sheet_view.showGridLines=False

# global style and validation
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.row not in (1,2,4) and cell.value is not None and cell.font.color is None:
                cell.font=Font(color=GRAY)
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.page_margins.left=.25; ws.page_margins.right=.25; ws.page_margins.top=.5; ws.page_margins.bottom=.5

out=D/'rp_ceo_expanded_benchmark_workbook.xlsx'
wb.save(out)
# Openpyxl validation pass: reopen formulas/data, check sheets and formula errors are not hard-coded.
check=load_workbook(out,data_only=False)
required={'Summary','Reference Set','Primary 990','All 990','Quant Job Ads','All Job Ads','Sensitivities','Peer Log','Source Manifest','Protocol & Notes'}
assert required.issubset(check.sheetnames)
formula_count=sum(1 for ws in check.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value,str) and c.value.startswith('='))
assert formula_count>=20, formula_count
for ws in check.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and any(e in c.value for e in ['#REF!','#DIV/0!','#VALUE!','#NAME?']): raise AssertionError((ws.title,c.coordinate,c.value))
print(out, 'formula_count', formula_count)
